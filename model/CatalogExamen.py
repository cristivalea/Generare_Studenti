import random
import openpyxl
from datetime import date
from Main import inscriere_facultate


def CatalogExamenNumeric(lista_studenti, disciplina, data_examen):
    note_sub_5 = 0.1
    note_egale_cu_5 = 0.05
    note_6 = 0.15
    note_7 = 0.25
    note_8 = 0.25
    note_9 = 0.15
    note_10 = 0.05

    nr_rand = random.randint(1, 70)
    prezenta = nr_rand / 100
    prezenta_examen = int(len(lista_studenti) * prezenta)

    studenti_sub_5 = int(note_sub_5 * prezenta_examen)
    studenti_5 = int(note_egale_cu_5 * prezenta_examen)
    studenti_6 = int(note_6 * prezenta_examen)
    studenti_7 = int(note_7 * prezenta_examen)
    studenti_8 = int(note_8 * prezenta_examen)
    studenti_9 = int(note_9 * prezenta_examen)
    studenti_10 = int(note_10 * prezenta_examen)

    absenti = prezenta_examen - (studenti_sub_5 + studenti_5 + studenti_6 + studenti_7 + studenti_8 + studenti_9 + studenti_10)

    kl5 = studenti_sub_5 + absenti
    k5 = studenti_5
    k6 = studenti_6
    k7 = studenti_7
    k8 = studenti_8
    k9 = studenti_9
    k10 = studenti_10

    lista_tuple = []
    for i in range(prezenta_examen):
        if i >= len(lista_studenti):
            break

        student = lista_studenti[i]

        if kl5 > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(), disciplina, data_examen, 4, 0, 5, 6, 10, 0, 14, 7)
            lista_tuple.append(tuplu)
            kl5 -= 1
        elif k5 > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(),disciplina, data_examen, 5, 0, 5, 6, 10, 0, 14, 7)
            lista_tuple.append(tuplu)
            k5 -= 1
        elif k6 > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(),disciplina, data_examen, 6, 0, 5, 6, 10, 0, 14, 7)
            lista_tuple.append(tuplu)
            k6 -= 1
        elif k7 > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(),disciplina, data_examen, 7, 0, 5, 6, 10, 0, 14, 7)
            lista_tuple.append(tuplu)
            k7 -= 1
        elif k8 > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(),disciplina, data_examen, 8, 0, 5, 6, 10, 0, 14, 7)
            lista_tuple.append(tuplu)
            k8 -= 1
        elif k9 > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(),disciplina, data_examen, 9, 0, 5, 6, 10, 0, 14, 7)
            lista_tuple.append(tuplu)
            k9 -= 1
        elif k10 > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(), disciplina, data_examen, 10, 0, 5, 6, 10, 0, 14, 7)
            lista_tuple.append(tuplu)
            k10 -= 1

    return lista_tuple

def  CatalogExamenCalificaticAR(lista_studenti, disciplina, data_examen):
    admis = 0.7
    respins = 0.3
    nr_rand = random.randint(1, 70)
    prezenta = nr_rand / 100
    prezenta_examen = int(len(lista_studenti) * prezenta)

    studenti_admisi = int(admis * prezenta_examen)
    stdenti_respinsi = int(respins * prezenta_examen)

    absente = prezenta_examen - (studenti_admisi + stdenti_respinsi)

    ka = studenti_admisi
    kr = stdenti_respinsi + absente

    lista_tuple = []
    for i in range(prezenta_examen + 1):
        if i >= len(lista_studenti):
            break
        student = lista_studenti[i]

        if ka > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(), disciplina, data_examen, "ADMIS", 7, 0, 7, 0)
            lista_tuple.append(tuplu)
            ka -= 1
        elif kr > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(), disciplina, data_examen, "RESPINS", 5, 0, 5, 0)
            lista_tuple.append(tuplu)
            kr -= 1
    return lista_tuple

def  CatalogExamenCalificatic(lista_studenti, disciplina, data_examen):
    insuficent = 0.1
    suficient = 0.3
    bine = 0.2
    foarte_bine = 0.2
    excelent = 0.2
    nr_rand = random.randint(1, 70)
    prezenta = nr_rand / 100
    prezenta_examen = int(len(lista_studenti) * prezenta)

    studenti_insuficent = int(insuficent * prezenta_examen)
    stdenti_suficient = int(suficient * prezenta_examen)
    studenti_bine = int(bine * prezenta_examen)
    studenti_foarte_bine = int(foarte_bine * prezenta_examen)
    studenti_excelent = int(excelent * prezenta_examen)

    absente = prezenta_examen - (studenti_insuficent + stdenti_suficient + studenti_bine + studenti_foarte_bine + studenti_excelent)

    ki = studenti_insuficent + absente
    ks = stdenti_suficient
    kb = studenti_bine
    kfb = studenti_foarte_bine
    ke = studenti_excelent

    lista_tuple = []
    for i in range(prezenta_examen + 1):
        if i >= len(lista_studenti):
            break
        student = lista_studenti[i]

        if ki > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(), disciplina, data_examen, "INSUFICIENT", 7, 0, 7, 0)
            lista_tuple.append(tuplu)
            ki -= 1
        elif ks > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(), disciplina, data_examen, "SUFICIENT", 5, 0, 5, 0)
            lista_tuple.append(tuplu)
            ks -= 1
        elif kb > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(), disciplina, data_examen, "BINE", 5, 0, 5, 0)
            lista_tuple.append(tuplu)
            kb -= 1
        elif kfb > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(), disciplina, data_examen, "FOARTE BINE", 5, 0, 5, 0)
            lista_tuple.append(tuplu)
            kfb -= 1
        elif ke > 0:
            tuplu = (student.get_numeFamilie(), student.get_prenume(), disciplina, data_examen, "EXCELENT", 5, 0, 5, 0)
            lista_tuple.append(tuplu)
            ke -= 1
    return lista_tuple


start = date(2024, 2, 1)
end = date(2024, 2, 3)
lista_studenti = inscriere_facultate(start, end)


rezultate = CatalogExamenNumeric(lista_studenti, "Programare", "2024-02-10")
rezultate1 = CatalogExamenCalificaticAR(lista_studenti, "Sport", "2023-12-15")
rezultate2 = CatalogExamenCalificatic(lista_studenti, "Comunicare", "2023-11-25")

def scrie_in_excel(lista_tuple, nume_fisier="catalog_examen_numeric.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalog Examen"

    ws.append(["Nume Student", "Prenume Student", "Disciplina", "Data Examen", "Nota Examen", "Nota Seminar", "Nota Laborator", "Nota Proiect", "Prezenta Examen", "Prezenta Seminar", "Prezenta Laborator", "Prezenta Proiect"])

    for tuplu in lista_tuple:
        ws.append(tuplu)

    wb.save(nume_fisier)
    print(f"Fișierul '{nume_fisier}' a fost creat cu succes!")

def scrie_in_excel_calificativ_ar(lista_tuple, nume_fisier="catalog_examen_calificativar.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalog Examen Calificativ"

    ws.append(["Nume Student", "Prenume Student", "Disciplina", "Data Examen", "Calificativ Examen Examen", "Prezenta Examen", "Prezenta Seminar", "Prezenta Laborator", "Prezenta Proiect"])

    for tuplu in lista_tuple:
        ws.append(tuplu)

    wb.save(nume_fisier)
    print(f"Fișierul '{nume_fisier}' a fost creat cu succes!")

def scrie_in_excel_calificativ(lista_tuple, nume_fisier="catalog_examen_calificativ.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalog Examen Calificativ"

    ws.append(["Nume Student", "Prenume Student", "Disciplina", "Data Examen", "Calificativ Examen", "Prezenta Examen", "Prezenta Seminar", "Prezenta Laborator", "Prezenta Proiect"])

    for tuplu in lista_tuple:
        ws.append(tuplu)

    wb.save(nume_fisier)
    print(f"Fișierul '{nume_fisier}' a fost creat cu succes!")


scrie_in_excel(rezultate)
scrie_in_excel_calificativ_ar(rezultate1)
scrie_in_excel_calificativ(rezultate2)